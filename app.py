"""
Camp Carysbrook Day-Off Scheduler — Streamlit App
"""

import io
import pandas as pd
import streamlit as st
import pulp
from collections import defaultdict

st.set_page_config(
    page_title="Camp Carysbrook · Staff Scheduler",
    page_icon="🌲",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Carysbrook Brand CSS ──────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Lato:wght@300;400;600&display=swap');

  html, body, [class*="css"] {
      font-family: 'Lato', sans-serif;
      background-color: #F5F0E8;
      color: #2E2E2E;
  }

  #MainMenu {visibility: hidden;}
  footer {visibility: hidden;}
  header {visibility: hidden;}

  /* ── Header banner ── */
  .cc-header {
      background: linear-gradient(135deg, #2D5A27 60%, #3D7035);
      padding: 1.6rem 2rem;
      border-radius: 10px;
      margin-bottom: 1.8rem;
      display: flex;
      align-items: center;
      gap: 1.4rem;
      box-shadow: 0 3px 12px rgba(45,90,39,0.25);
  }
  .cc-header-text h1 {
      font-family: 'Playfair Display', serif;
      color: #FFFFFF;
      font-size: 1.9rem;
      margin: 0;
      line-height: 1.2;
  }
  .cc-header-text p {
      color: #C8DFC5;
      font-size: 0.85rem;
      margin: 0.25rem 0 0 0;
      font-weight: 300;
      letter-spacing: 0.06em;
      text-transform: uppercase;
  }
  .cc-icon { font-size: 2.8rem; }

  /* ── Headings ── */
  h2, h3 {
      font-family: 'Playfair Display', serif !important;
      color: #5C3317 !important;
  }

  /* ── Primary button ── */
  .stButton > button[kind="primary"] {
      background-color: #5C3317 !important;
      border: none !important;
      color: white !important;
      font-family: 'Lato', sans-serif !important;
      font-weight: 600 !important;
      letter-spacing: 0.06em !important;
      border-radius: 6px !important;
      padding: 0.65rem 1.6rem !important;
      transition: background-color 0.2s ease !important;
  }
  .stButton > button[kind="primary"]:hover {
      background-color: #7A4420 !important;
  }

  /* ── Download buttons ── */
  .stDownloadButton > button {
      background-color: #FFFFFF !important;
      border: 1.5px solid #2D5A27 !important;
      color: #2D5A27 !important;
      font-family: 'Lato', sans-serif !important;
      font-weight: 600 !important;
      border-radius: 6px !important;
      transition: all 0.2s ease !important;
  }
  .stDownloadButton > button:hover {
      background-color: #2D5A27 !important;
      color: white !important;
  }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {
      background-color: #2D5A27 !important;
  }
  [data-testid="stSidebar"] * {
      color: #E8F0E7 !important;
  }
  [data-testid="stSidebar"] h1,
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3 {
      color: #FFFFFF !important;
      font-family: 'Playfair Display', serif !important;
  }
  [data-testid="stSidebar"] hr {
      border-color: #4A7A43 !important;
  }
  [data-testid="stSidebar"] code {
      background-color: #1F3F1C !important;
      color: #C8DFC5 !important;
      border-radius: 4px;
      padding: 0.2rem 0.4rem;
  }
  [data-testid="stSidebar"] .stDownloadButton > button {
      background-color: #5C3317 !important;
      border: none !important;
      color: white !important;
  }
  [data-testid="stSidebar"] .stDownloadButton > button:hover {
      background-color: #7A4420 !important;
  }

  /* ── File uploader ── */
  [data-testid="stFileUploader"] {
      background-color: #FFFFFF;
      border: 2px dashed #2D5A27 !important;
      border-radius: 8px;
  }

  /* ── Expander ── */
  [data-testid="stExpander"] {
      border: 1px solid #C8DFC5 !important;
      border-radius: 8px !important;
      background-color: #FFFFFF;
  }

  /* ── Alerts ── */
  [data-testid="stAlert"] {
      border-radius: 6px !important;
  }

  /* ── Cards / metric boxes ── */
  [data-testid="stMetric"] {
      background-color: #FFFFFF;
      border: 1px solid #C8DFC5;
      border-left: 4px solid #5C3317;
      border-radius: 8px;
      padding: 0.8rem 1rem;
  }

  /* ── Footer ── */
  .cc-footer {
      margin-top: 3rem;
      padding-top: 1rem;
      border-top: 1px solid #C8B89A;
      text-align: center;
      color: #9A7B5A;
      font-size: 0.78rem;
      letter-spacing: 0.05em;
  }

  /* ── Bar chart override colour ── */
  [data-testid="stVegaLiteChart"] canvas { border-radius: 6px; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="cc-header">
  <div class="cc-icon">🌲</div>
  <div class="cc-header-text">
    <h1>Camp Carysbrook · Staff Scheduler</h1>
    <p>Days-off scheduling &nbsp;·&nbsp; Est. 1923 &nbsp;·&nbsp; Riner, Virginia</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Colour palette for days — green → brown tones ─────────────────────────────
DAY_COLOURS = [
    "#2D5A27", "#5C3317", "#6B9E64", "#7A4420", "#4A7A43",
    "#9A6040", "#8FB887", "#C8936A", "#355E2E", "#B87040",
    "#A3C49B", "#D4956A", "#527A4C", "#E8B088",
]

REQUIRED_COLS = {"Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age"}

TEMPLATE_DATA = [
    ("Counselor 19","Lark","RRG","Archery","Riflery",25),
    ("Counselor 22","Mockingbird","","Dock","",25),
    ("Counselor 3","Sparrow","","Barn","",23),
    ("Counselor 10","Swan","","OLS","",22),
    ("Counselor 12","Hummingbird","AHC","Riflery","",21),
    ("Counselor 2","Cardinal","","Fencing","",21),
    ("Counselor 20","Sparrow","","Dock","",21),
    ("Counselor 21","Bobwhite","","Marketing","",21),
    ("Counselor 33","Swallow","AAC","Trips","",21),
    ("Counselor 16","Finch","","Barn","",20),
    ("Counselor 24","Cardinal","","Dock","",20),
    ("Counselor 27","Mallard","AHC","Dock","",20),
    ("Counselor 34","Wren","","A&C","Riflery",20),
    ("Counselor 1","Bobwhite","","A&C","",17),
    ("Counselor 13","Lark","","Nature","",19),
    ("Counselor 15","Finch","","Dock","Rec Sports",19),
    ("Counselor 23","Wren","","Store","Expeditions",19),
    ("Counselor 29","Swallow","","Barn","",19),
    ("Counselor 30","Mockingbird","AHC","Drama","",19),
    ("Counselor 32","Bluebird","G2G","Tumbling","",19),
    ("Counselor 5","Bluebird","","Climbing","",19),
    ("Counselor 6","Dove","","Riflery","Synchro",19),
    ("Counselor 9","Dove","AAC","Rec Sports","",19),
    ("Counselor 28","Oriole","AHC","Dance","",18),
    ("Counselor 31","Swan","","Canoeing","",18),
    ("Counselor 4","Nurse","","Nurse","",18),
    ("Counselor 7","Oriole","","Riflery","",18),
    ("Counselor 26","Mallard","","Archery","",17),
    ("Counselor 8","Chick","HC","Performances","",17),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalise(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["Activity1", "Activity2"]:
        df[col] = df[col].fillna("").str.strip().str.lower()
    for col in ["Cabin", "Leadership", "Name"]:
        df[col] = df[col].fillna("").str.strip()
    df["Age"] = pd.to_numeric(df["Age"], errors="coerce").fillna(0).astype(int)
    return df


def schedule(df: pd.DataFrame, num_days: int, days_per: int) -> pd.DataFrame:
    counselors = df["Name"].tolist()
    days = list(range(1, num_days + 1))
    n = len(counselors)

    prob = pulp.LpProblem("CampDayOff", pulp.LpMinimize)
    x = {(i, d): pulp.LpVariable(f"x_{i}_{d}", cat="Binary") for i in range(n) for d in days}

    for i in range(n):
        prob += pulp.lpSum(x[i, d] for d in days) == days_per

    cabin_groups = defaultdict(list)
    for i, row in df.iterrows():
        if row["Cabin"]:
            cabin_groups[row["Cabin"]].append(i)
    for cabin, members in cabin_groups.items():
        if len(members) > 1:
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= 1

    act_groups = defaultdict(set)
    for i, row in df.iterrows():
        for col in ["Activity1", "Activity2"]:
            if row[col]:
                act_groups[row[col]].add(i)
    for act, members in act_groups.items():
        members = list(members)
        if len(members) > 1:
            cap = max(1, len(members) // 2)
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= cap

    lead_groups = defaultdict(list)
    for i, row in df.iterrows():
        if row["Leadership"]:
            lead_groups[row["Leadership"]].append(i)
    for role, members in lead_groups.items():
        if len(members) > 1:
            cap = max(1, len(members) // 2)
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= cap

    ages = df["Age"].tolist()
    max_age = max(ages) if ages else 0
    seniors = [i for i in range(n) if ages[i] >= 20]
    if seniors:
        peak = pulp.LpVariable("peak_seniors", lowBound=0)
        for d in days:
            prob += pulp.lpSum(x[i, d] for i in seniors) <= peak
        prob += peak

    solver = pulp.PULP_CBC_CMD(msg=0, timeLimit=60)
    prob.solve(solver)

    status = pulp.LpStatus[prob.status]
    if status not in ("Optimal", "Feasible"):
        raise RuntimeError(
            f"No valid schedule found (solver status: {status}). "
            "Try increasing the number of available days."
        )

    assignment = {}
    for i in range(n):
        for d in days:
            if round(pulp.value(x[i, d]) or 0) == 1:
                assignment[counselors[i]] = d
                break

    result = df.copy()
    result["Day Off"] = result["Name"].map(assignment)
    return result.sort_values("Day Off").reset_index(drop=True)


def validate(result: pd.DataFrame):
    issues = []
    for (cabin, day), grp in result[result["Cabin"] != ""].groupby(["Cabin", "Day Off"]):
        if len(grp) > 1:
            issues.append(f"Cabin **{cabin}** — day {day}: {', '.join(grp['Name'])}")
    act_totals = defaultdict(set)
    act_day = defaultdict(lambda: defaultdict(list))
    for _, row in result.iterrows():
        for col in ["Activity1", "Activity2"]:
            if row[col]:
                act_totals[row[col]].add(row["Name"])
                act_day[row[col]][row["Day Off"]].append(row["Name"])
    for act, day_map in act_day.items():
        total = len(act_totals[act])
        if total <= 1:
            continue
        for day, names in day_map.items():
            if len(names) > total / 2:
                issues.append(f"Activity **{act}** — {len(names)}/{total} off on day {day}")
    lead_totals = defaultdict(set)
    lead_day = defaultdict(lambda: defaultdict(list))
    for _, row in result.iterrows():
        if row["Leadership"]:
            lead_totals[row["Leadership"]].add(row["Name"])
            lead_day[row["Leadership"]][row["Day Off"]].append(row["Name"])
    for role, day_map in lead_day.items():
        total = len(lead_totals[role])
        if total <= 1:
            continue
        for day, names in day_map.items():
            if len(names) > total / 2:
                issues.append(f"Leadership **{role}** — {len(names)}/{total} off on day {day}")
    return issues


def colour_row(row, palette):
    day = row["Day Off"]
    if pd.isna(day):
        return [""] * len(row)
    colour = palette.get(int(day), "#FFFFFF")
    r = int(colour[1:3], 16)
    g = int(colour[3:5], 16)
    b = int(colour[5:7], 16)
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    text = "black" if brightness > 128 else "white"
    return [f"background-color:{colour};color:{text}"] * len(row)


def to_excel(df: pd.DataFrame) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day-Off Schedule"

    days = sorted(df["Day Off"].dropna().unique())
    palette = {int(d): DAY_COLOURS[i % len(DAY_COLOURS)] for i, d in enumerate(days)}

    headers = ["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age", "Day Off"]
    header_fill = PatternFill("solid", fgColor="2D5A27")
    header_font = Font(bold=True, color="FFFFFF", name="Georgia", size=11)
    thin = Side(style="thin", color="C8B89A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, (_, row) in enumerate(df[headers].iterrows(), 2):
        day_val = row["Day Off"]
        hex_colour = palette.get(int(day_val), "FFFFFF").lstrip("#") if not pd.isna(day_val) else "F5F0E8"
        r = int(hex_colour[0:2], 16)
        g = int(hex_colour[2:4], 16)
        b = int(hex_colour[4:6], 16)
        brightness = (r * 299 + g * 587 + b * 114) / 1000
        text_colour = "000000" if brightness > 128 else "FFFFFF"
        fill = PatternFill("solid", fgColor=hex_colour)
        font = Font(name="Calibri", size=10, color=text_colour)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    col_widths = [22, 16, 14, 16, 16, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    num_days = st.slider(
        "Days in schedule window",
        min_value=3, max_value=7, value=5,
        help="How many days the schedule spans (e.g. 5 for one week)."
    )
    days_per = st.slider("Days off per counselor", min_value=1, max_value=3, value=1)

    st.divider()

    st.markdown("## 📋 Template")
    st.write("Download the template to see the required column format:")
    template_df = pd.DataFrame(
        TEMPLATE_DATA,
        columns=["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age"]
    )
    template_bytes = io.BytesIO()
    template_df.to_csv(template_bytes, index=False)
    st.download_button(
        "⬇️ Download template CSV",
        template_bytes.getvalue(),
        file_name="carysbrook_counselor_template.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.divider()

    st.markdown("**Required columns:**")
    st.code("Name, Cabin, Leadership,\nActivity1, Activity2, Age")
    st.caption("Leave cells blank where not applicable.")


# ── Main ──────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload your counselor file (.csv or .xlsx)",
    type=["csv", "xlsx"],
    help="File must include the columns listed in the sidebar."
)

df_raw = None
if uploaded:
    try:
        if uploaded.name.endswith(".xlsx"):
            df_raw = pd.read_excel(uploaded)
        else:
            df_raw = pd.read_csv(uploaded)
        missing = REQUIRED_COLS - set(df_raw.columns)
        if missing:
            st.error(f"Missing columns: {', '.join(sorted(missing))}. Please check your file matches the template.")
            df_raw = None
        else:
            col1, col2, col3 = st.columns(3)
            col1.metric("Counselors", len(df_raw))
            col2.metric("Cabins", df_raw["Cabin"].replace("", pd.NA).dropna().nunique())
            col3.metric("Activities", pd.concat([df_raw["Activity1"], df_raw["Activity2"]]).replace("", pd.NA).dropna().nunique())
            st.success(f"✅ **{uploaded.name}** loaded successfully.")
    except Exception as e:
        st.error(f"Could not read file: {e}")
else:
    st.info("👆 Upload your counselor spreadsheet above, or download the template from the sidebar to get started.")

if df_raw is not None:
    with st.expander("Preview uploaded data", expanded=False):
        st.dataframe(df_raw, use_container_width=True)

    st.write("")
    if st.button("🗓️ Generate Schedule", type="primary", use_container_width=True):
        df = normalise(df_raw)
        with st.spinner("Finding the best schedule…"):
            try:
                result = schedule(df, num_days=num_days, days_per=days_per)
            except RuntimeError as e:
                st.error(str(e))
                st.stop()

        st.success("Schedule generated!")

        issues = validate(result)
        if issues:
            with st.expander("⚠️ Constraint warnings", expanded=True):
                for issue in issues:
                    st.warning(issue)
        else:
            st.info("✅ All cabin, activity, and leadership constraints satisfied.")

        days_sorted = sorted(result["Day Off"].dropna().unique())
        palette = {int(d): DAY_COLOURS[i % len(DAY_COLOURS)] for i, d in enumerate(days_sorted)}

        st.markdown("### 📊 Counselors off per day")
        counts = result.groupby("Day Off")["Name"].count().reset_index()
        counts.columns = ["Day", "Counselors Off"]
        st.bar_chart(counts.set_index("Day"), color="#2D5A27")

        st.markdown("### 📅 Full Schedule")
        display_cols = ["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age", "Day Off"]
        styled = (
            result[display_cols]
            .style
            .apply(colour_row, palette=palette, axis=1)
            .set_properties(**{"font-size": "13px"})
        )
        st.dataframe(styled, use_container_width=True, height=600)

        st.markdown("### ⬇️ Download")
        col1, col2 = st.columns(2)
        with col1:
            csv_bytes = result[display_cols].to_csv(index=False).encode()
            st.download_button(
                "Download as CSV",
                csv_bytes,
                file_name="carysbrook_day_off_schedule.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with col2:
            try:
                xl_bytes = to_excel(result)
                st.download_button(
                    "Download as Excel (colour-coded)",
                    xl_bytes,
                    file_name="carysbrook_day_off_schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except ImportError:
                st.warning("Install `openpyxl` for Excel export.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="cc-footer">
  Camp Carysbrook &nbsp;·&nbsp; 3500 Camp Carysbrook Road, Riner, VA 24149 &nbsp;·&nbsp; Est. 1923
</div>
""", unsafe_allow_html=True)
